import streamlit as st
import pandas as pd
import zipfile

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from num2words import num2words
from datetime import datetime

# =========================
# PAGE CONFIG
# =========================

st.set_page_config(
    page_title="Estimate Generator",
    layout="wide"
)

st.title("Estimate Generator")

# =========================
# MASTER CONFIG
# =========================

MASTER_SHEET_ID = (
    "1kjkKVIJeYUNLs_uO0SthMSJ6-7GNKd80K4pPX7nBeig"
)

# =========================
# COMMON STYLES
# =========================

thin = Side(style="thin", color="000000")

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

header_fill = PatternFill(
    start_color="FFA500",
    end_color="FFA500",
    fill_type="solid"
)

grey_fill = PatternFill(
    start_color="D9D9D9",
    end_color="D9D9D9",
    fill_type="solid"
)

blue_fill = PatternFill(
    start_color="0070C0",
    end_color="0070C0",
    fill_type="solid"
)

bold_font = Font(
    bold=True,
    name="Gill Sans MT",
    size=10
)

normal_font = Font(
    name="Gill Sans MT",
    size=10
)

title_font = Font(
    bold=True,
    size=18,
    color="0070C0",
    name="Gill Sans MT"
)

white_font = Font(
    bold=True,
    color="FFFFFF",
    name="Gill Sans MT",
    size=10
)

center_alignment = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

left_alignment = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True
)

# =========================
# LOAD GOOGLE SHEET TAB
# =========================

def load_sheet(sheet_name):

    url = (
        f"https://docs.google.com/spreadsheets/d/"
        f"{MASTER_SHEET_ID}/gviz/tq?"
        f"tqx=out:csv&sheet={sheet_name}"
    )

    df = pd.read_csv(
        url,
        engine="python"
    )

    df.columns = (
        df.columns
        .str.strip()
    )

    return df

# =========================
# LOAD MASTER TABLES
# =========================

try:

    client_master_df = load_sheet(
        "ClientMaster"
    )

    brand_master_df = load_sheet(
        "BrandMaster"
    )

    rate_master_df = load_sheet(
        "RateMaster"
    )

    st.success(
        "Master tables loaded successfully"
    )

except Exception as e:

    st.error(
        f"Error loading master tables: {e}"
    )

# =========================
# APPLY CELL STYLE
# =========================

def style_range(
    ws,
    cell_range,
    fill=None,
    font=None,
    alignment=None,
    left=None,
    right=None,
    top=None,
    bottom=None

):
    border_style = Border(

            left=Side(style="thin", color="000000") if left else Side(style=None),

            right=Side(style="thin", color="000000") if right else Side(style=None),

            top=Side(style="thin", color="000000") if top else Side(style=None),

            bottom=Side(style="thin", color="000000") if bottom else Side(style=None)

        )

    rows = ws[cell_range]

    for row in rows:

        for cell in row:

            cell.border = border_style

            if fill:
                cell.fill = fill

            if font:
                cell.font = font

            if alignment:
                cell.alignment = alignment


# =========================
# HEADER CREATION
# =========================

orange_font = Font(
    bold=False,
    color="FFA500",
    name="Gill Sans MT",
    size=10
)

def create_header(
    ws,
    client,
    address,
    pan,
    gstin,
    brand_manager,
    estimate_name
):

    ws.merge_cells("B2:I2")

    ws["B2"] = "ANNEXURE OF COMMERCIAL TERMS"

    style_range(
        ws,
        "B2:I2",
        fill=blue_fill,
        font=white_font,
        alignment=center_alignment,
        left=True,
        top=True,
        right=True,
        bottom=True

    )

    ws.merge_cells("B3:G4")

    ws["B3"] = "ENCEPT PREMEDIA PRIVATE LIMITED"

    style_range(
        ws,
        "B3:G4",
        font=title_font,
        alignment=center_alignment,
        left=True,
        top=True,
        right=True,
        bottom=False
    )

    ws.merge_cells("B5:G5")

    ws["B5"] = "PACKAGING BRAND MANAGEMENT"

    style_range(
        ws,
        "B5:G5",
        font=orange_font,
        alignment=center_alignment,
        left=True,
        top=False,
        right=True,
        bottom=True
        
    )

    ws.merge_cells("H3:I6")
    style_range(
        ws,
        "H3:I6",
        font=orange_font,
        alignment=center_alignment,
        left=True,
        top=False,
        right=True,
        bottom=True
        
    )

    ws.merge_cells("B6:G6")

    ws["B6"] = (
        "C-111-118, BSEL Tech Park, "
        "Vashi, Navi Mumbai-400703"
    )

    style_range(
        ws,
        "B6:G6",
        fill=grey_fill,
        font=normal_font,
        alignment=center_alignment,
        left=True,
        top=True,
        right=True,
        bottom=True
    )

    ws.merge_cells("B7:I7")

    style_range(
        ws,
        "B7:I7",
        left=True,
        top=True,
        right=True,
        bottom=True
    )

    ws.merge_cells("B8:E8")

    ws["B8"] = (
        f"Kind Attn.: {brand_manager}"
    )

    style_range(
        ws,
        "B8:E8",
        fill=grey_fill,
        font=bold_font,
        alignment=left_alignment,
        left=True,
        top=True,
        right=True,
        bottom=True
    )

    ws.merge_cells("B9:E13")

    ws["B9"] = (
        f"{client}\n"
        f"{address}\n"
        f"PAN No.: {pan}\n"
        f"GSTIN: {gstin}"
    )

    style_range(
        ws,
        "B9:E13",
        font=normal_font,
        alignment=left_alignment,
        left=True,
        top=True,
        right=True,
        bottom=True
    )

    ws.merge_cells("F8:I8")

    ws["F8"] = "Estimate No."

    style_range(
        ws,
        "F8:I8",
        fill=grey_fill,
        font=bold_font,
        alignment=left_alignment,
        left=True,
        top=True,
        right=True,
        bottom=True
    )

    ws.merge_cells("F9:I10")

    ws["F9"] = estimate_name

    style_range(
        ws,
        "F9:I10",
        font=normal_font,
        alignment=left_alignment,
        left=True,
        top=True,
        right=True,
        bottom=True
    )

    ws.merge_cells("F11:I11")

    ws["F11"] = "Date"

    style_range(
        ws,
        "F11:I11",
        fill=grey_fill,
        font=bold_font,
        alignment=left_alignment,
        left=True,
        top=True,
        right=True,
        bottom=True
    )

    ws.merge_cells("F12:I13")

    ws["F12"] = (
        datetime.now()
        .strftime("%d %b %Y")
    )

    style_range(
        ws,
        "F12:I13",
        font=normal_font,
        alignment=left_alignment,
        left=True,
        top=True,
        right=True,
        bottom=True
    )

# =========================
# TABLE HEADER
# =========================

    # =========================
    # PERIOD LINE
    # =========================

    period_line = (
        f"Period: "
        f"{start_date.strftime('%d-%b-%Y')} "
        f"to "
        f"{end_date.strftime('%d-%b-%Y')}"
    )

    ws.merge_cells("B15:I15")

    ws["B15"] = period_line

    style_range(
        ws,
        "B15:I15",
        fill=grey_fill,
        font=bold_font,
        alignment=center_alignment,
        left = True,
        right = True,
        top = True,
        bottom = False
    )

    # =========================
    # SERVICE TITLE
    # =========================

    ws.merge_cells("B16:I16")

    ws["B16"] = (
        "Artwork Production Services"
    )

    ws["B16"].font = bold_font
    ws["B16"].alignment = center_alignment
    ws["B16"].border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style=None)
    )


    create_table_header(ws)
    # =========================
    # LOGO INSERTION
    # =========================

    try:

        logo = Image("encept.png")

        logo.width = 170
        logo.height = 55

        ws.add_image(
            logo,
            "H3"
        )

    except Exception as e:

        print(
            f"Logo insert failed: {e}"
        )
    
# =========================
# COLUMN WIDTHS
# =========================

def set_column_widths(ws):

    widths = {

        "B": 10,
        "C": 28,
        "D": 16,
        "E": 60,
        "F": 14,
        "G": 10,
        "H": 12,
        "I": 14,
        "J": 50

    }

    for col, width in widths.items():

        ws.column_dimensions[
            col
        ].width = width




# =========================
# CALCULATE ESTIMATE
# =========================

def calculate_estimate(
    brand_df,
    client,
    rate_master_df
):

    summary_rows = []

    grouped = (
        brand_df.groupby("Task_Name")
        .size()
        .reset_index(name="Count")
    )

    rate_row = rate_master_df[
        rate_master_df["Client Name"] == client
    ]

    for _, row in grouped.iterrows():

        task_name = str(
            row["Task_Name"]
        ).strip()

        count = row["Count"]

        rate = 0

        if (
            not rate_row.empty and
            task_name in rate_row.columns
        ):

            try:

                rate = float(
                    rate_row.iloc[0][task_name]
                )

            except:

                rate = 0

        amount = count * rate

        summary_rows.append({

            "Task_Name": task_name,
            "Count": count,
            "Rate": rate,
            "Amount": amount

        })

    return pd.DataFrame(summary_rows)



# =========================
# TABLE HEADER
# =========================

def create_table_header(ws):

    headers = [
        "Sr. No",
        "Scope",
        "Date",
        "Description",
        "SAC No",
        "Qty",
        "Rate",
        "Amount"
    ]

    row = 14

    for col_index, value in enumerate(headers, start=2):

        cell = ws.cell(
            row=row,
            column=col_index,
            value=value
        )

        cell.fill = header_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = center_alignment


# =========================
# GENERATE EXCEL
# =========================

def generate_estimate_excel(
    sequence_no,
    brand,
    brand_df,
    estimate_df,
    client,
    client_master_df,
    brand_master_df
):

    wb = Workbook()

    ws = wb.active

    sequence_str = str(
        sequence_no
    ).zfill(2)

    ws.title = (
        f"{sequence_str}_{brand}"
    )[:31]

    set_column_widths(ws)

    # =========================
    # CLIENT DETAILS
    # =========================

    client_row = client_master_df[
        client_master_df["Client Name"] == client
    ].iloc[0]

    address = client_row["Address"]
    pan = client_row["PAN"]
    gstin = client_row["GSTIN"]
    shortcode = client_row["ShortCode"]
    tax_percent = float(
        client_row["TaxPercent"]
    )

    brand_row = brand_master_df[

        (
            brand_master_df["Client_Name"]
            == client
        )

        &

        (
            brand_master_df["Brand_Name"]
            == brand
        )

    ]

    brand_manager = ""

    if not brand_row.empty:

        brand_manager = (
            brand_row.iloc[0]
            ["Brand_Manager_Name"]
        )

    current_month = datetime.now().strftime(
        "%B_%Y"
    )

    estimate_name = (

        f"{sequence_str}_"
        f"EPPL_"
        f"{shortcode}_"
        f"{brand}_"
        f"{current_month}_"
        f"{brand_manager}"

    )

    create_header(

        ws,
        client,
        address,
        pan,
        gstin,
        brand_manager,
        estimate_name

    )

    
    # =========================
    # PERIOD LINE
    # =========================

    period_line = (
        f"Period: "
        f"{start_date.strftime('%d-%b-%Y')} "
        f"to "
        f"{end_date.strftime('%d-%b-%Y')}"
    )

    ws.merge_cells("B15:I15")

    ws["B15"] = period_line

    style_range(
        ws,
        "B15:I15",
        fill=grey_fill,
        font=bold_font,
        alignment=center_alignment
    )

    # =========================
    # SERVICE TITLE
    # =========================

    ws.merge_cells("B16:I16")

    ws["B16"] = (
        "Artwork Production Services"
    )

    ws["B16"].font = bold_font
    ws["B16"].alignment = center_alignment

    style_range(
        ws,
        "B17:I17",
        left=True,
        right=True,
        top=False,
        bottom=False
    )


    # =========================
    # LOGO INSERTION
    # =========================

    try:

        logo = Image("encept.png")

        logo.width = 170
        logo.height = 55

        ws.add_image(
            logo,
            "H3"
        )

    except Exception as e:

        print(
            f"Logo insert failed: {e}"
        )


    # =========================
    # BUILD ROWS
    # =========================

    output_rows = []

    serial_no = 1

    for _, est_row in estimate_df.iterrows():

        task_name = est_row["Task_Name"]

        qty = est_row["Count"]

        rate = est_row["Rate"]

        amount = est_row["Amount"]

        task_rows = brand_df[
            brand_df["Task_Name"]
            == task_name
        ]

        task_rows = task_rows.sort_values(
            by="AMT_Inititate_Time"
        )

        # SUMMARY ROW
        # Amount column (I) will use Excel formula
        # instead of static value

        output_rows.append([

            serial_no,
            "CBB" if task_name in [
                "Artwork-Secondary Packaging(CLD)",
                "Artwork - Secondary Packaging (CLD)"
            ] else task_name,
            "",
            "",
            "998391",
            qty,
            rate,
            "",     # Formula added later
            ""

        ])

        # DETAIL ROWS

        for _, task_row in task_rows.iterrows():

            date_value = ""

            if pd.notna(
                task_row["AMT_Inititate_Time"]
            ):

                date_value = pd.to_datetime(
                    task_row["AMT_Inititate_Time"]
                ).strftime("%d-%b-%Y")

            description = (
                str(task_row["Variant_Name"]).title() + " | " +
                str(task_row["SKU_Name"]).title() + " | " +
                str(task_row["Package_Format"]).title()
            )            
            

            brief = ""

            if "Brief" in task_row:

                if pd.notna(
                    task_row["Brief"]
                ):

                    brief = str(
                        task_row["Brief"]
                    )

            output_rows.append([

                "",
                "",
                date_value,
                description,
                "",
                "",
                "",
                "",
                brief

            ])

        output_rows.append(
            [""] * 9
        )

        serial_no += 1

    # =========================
    # WRITE BODY
    # =========================

    start_row = 18

    for row_data in output_rows:

        for col_offset, value in enumerate(
            row_data,
            start=2
        ):

            # Column I (Amount)
            # Formula = Qty × Rate
            if col_offset == 9 and row_data[0] != "":

                value = f"=ROUND(G{start_row}*H{start_row},0)"

            cell = ws.cell(
                row=start_row,
                column=col_offset,
                value=value
            )

            cell.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style=None),
                bottom=Side(style=None)
            )

            white_fill = PatternFill(
                start_color="FFFFFF",
                end_color="FFFFFF",
                fill_type="solid"
            )            

            cell.fill = white_fill
            cell.font = normal_font

            # Comma separator formatting
            if col_offset == 9:
                cell.number_format = '#,##0'

            if col_offset in [2,3,4,5,6,7,8,9]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

        ws.row_dimensions[
            start_row
        ].height = 22

        start_row += 1

    # =========================
    # TOTALS
    # =========================

    # =========================
    # FORMULA-BASED TOTALS
    # =========================

    first_data_row = 18
    last_data_row = start_row - 1

    totals_row = start_row + 1

    ws.merge_cells(
        start_row=totals_row,
        start_column=6,
        end_row=totals_row,
        end_column=8
    )

    ws.cell(
        row=totals_row,
        column=6,
        value="Total"
    )

    ws.cell(
        row=totals_row,
        column=9,
        value=f"=ROUND(SUM(I{first_data_row}:I{last_data_row}),0)"
    )

    totals_row += 1

    ws.merge_cells(
        start_row=totals_row,
        start_column=6,
        end_row=totals_row,
        end_column=8
    )

    ws.cell(
        row=totals_row,
        column=6,
        value=f"GST @ {tax_percent}%"
    )

    ws.cell(
        row=totals_row,
        column=9,
        value=f"=ROUND(I{totals_row-1}*{tax_percent}/100,0)"
    )

    totals_row += 1

    ws.merge_cells(
        start_row=totals_row,
        start_column=6,
        end_row=totals_row,
        end_column=8
    )

    ws.cell(
        row=totals_row,
        column=6,
        value="Grand Total (Including Tax)"
    )

    ws.cell(
        row=totals_row,
        column=9,
        value=f"=ROUND(I{totals_row-2}+I{totals_row-1},0)"
    )

    for r in range(
        totals_row - 2,
        totals_row + 1
    ):

        for c in range(6, 10):

            cell = ws.cell(
                row=r,
                column=c
            )

            cell.border = border
            cell.font = bold_font
            cell.alignment = center_alignment

            # Comma separator for totals
            if c == 9:
                cell.number_format = '#,##0' 

    # =========================
    # AMOUNT IN WORDS
    # =========================

    # =========================
    # AMOUNT IN WORDS
    # =========================

    amount_words_formula = f"""=IF(I{totals_row}=0,"Rupees Zero Only",
"Rupees " &
IF(INT(I{totals_row}/100000)>0,
  IF(INT(I{totals_row}/100000)<20,
    INDEX({{"","One","Two","Three","Four","Five","Six","Seven","Eight","Nine","Ten","Eleven","Twelve","Thirteen","Fourteen","Fifteen","Sixteen","Seventeen","Eighteen","Nineteen"}},1,1+INT(I{totals_row}/100000)),
    INDEX({{"","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"}},1,1+INT(INT(I{totals_row}/100000)/10)) &
    IF(MOD(INT(I{totals_row}/100000),10)>0," "&INDEX({{"","One","Two","Three","Four","Five","Six","Seven","Eight","Nine"}},1,1+MOD(INT(I{totals_row}/100000),10)),"")
  ) & " Lakh ",
  ""
) &
IF(MOD(INT(I{totals_row}/1000),100)>0,
  IF(MOD(INT(I{totals_row}/1000),100)<20,
    INDEX({{"","One","Two","Three","Four","Five","Six","Seven","Eight","Nine","Ten","Eleven","Twelve","Thirteen","Fourteen","Fifteen","Sixteen","Seventeen","Eighteen","Nineteen"}},1,1+MOD(INT(I{totals_row}/1000),100)),
    INDEX({{"","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"}},1,1+INT(MOD(INT(I{totals_row}/1000),100)/10)) &
    IF(MOD(MOD(INT(I{totals_row}/1000),100),10)>0," "&INDEX({{"","One","Two","Three","Four","Five","Six","Seven","Eight","Nine"}},1,1+MOD(MOD(INT(I{totals_row}/1000),100),10)),"")
  ) & " Thousand ",
  ""
) &
IF(MOD(INT(I{totals_row}/100),10)>0,
  INDEX({{"","One","Two","Three","Four","Five","Six","Seven","Eight","Nine"}},1,1+MOD(INT(I{totals_row}/100),10)) & " Hundred ",
  ""
) &
IF(MOD(I{totals_row},100)>0,
  "and " &
  IF(MOD(I{totals_row},100)<20,
    INDEX({{"","One","Two","Three","Four","Five","Six","Seven","Eight","Nine","Ten","Eleven","Twelve","Thirteen","Fourteen","Fifteen","Sixteen","Seventeen","Eighteen","Nineteen"}},1,1+MOD(I{totals_row},100)),
    INDEX({{"","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"}},1,1+INT(MOD(I{totals_row},100)/10)) &
    IF(MOD(MOD(I{totals_row},100),10)>0," "&INDEX({{"","One","Two","Three","Four","Five","Six","Seven","Eight","Nine"}},1,1+MOD(MOD(I{totals_row},100),10)),"")
  ),
  ""
) & " Only")"""

    words_row = totals_row + 1

    white_fill = PatternFill(
        start_color="FFFFFF",
        end_color="FFFFFF",
        fill_type="solid"
    )

    ws.merge_cells(
        start_row=words_row,
        start_column=2,
        end_row=words_row,
        end_column=9
    )

    ws.cell(
        row=words_row,
        column=2,
        value=f'="Amount in words : " & {amount_words_formula[1:]}'

    )

    style_range(
        ws,
        f"B{words_row}:I{words_row}",
        fill=white_fill,
        font=bold_font,
        alignment=left_alignment,
        left=True,
        right=True,
        top=False,
        bottom=True
    )

    blank_row = words_row + 1

    ws.merge_cells(
        f"B{blank_row}:I{blank_row}"
    )

    style_range(
        ws,
        f"B{blank_row}:I{blank_row}",
        fill=grey_fill,
        left=True,
        right=True,
        top=True,
        bottom=True
    )    

    # =========================
    # TERMS & CONDITIONS
    # =========================

    terms_row = totals_row - 2

    ws.merge_cells(
        f"B{terms_row}:E{terms_row}"
    )

    ws[f"B{terms_row}"] = "Terms & Conditions"

    style_range(
        ws,
        f"B{terms_row}:E{terms_row}",
        fill=grey_fill,
        font=bold_font,
        alignment=left_alignment,
        left=True,
        right=True,
        top=True,
        bottom=True
    )

    ws.merge_cells(
        f"B{terms_row+1}:E{terms_row+2}"
    )

    ws[f"B{terms_row+1}"] = (
        "Payment Terms: 30 days from the date of invoice.\n"
        "Delivery: As per agreement."
    )

    style_range(
        ws,
        f"B{terms_row+1}:E{terms_row+2}",
        font=normal_font,
        alignment=left_alignment,
        left=True,
        right=True,
        top=False,
        bottom=True
    )


    # =========================
    # FOOTER
    # =========================

    footer_row = words_row + 2

    ws.merge_cells(
        start_row=footer_row,
        start_column=2,
        end_row=footer_row + 4,
        end_column=5
    )

    ws.cell(
        row=footer_row,
        column=2,
        value=(
            "GSTIN: 27AADCE1404G1Z5\n"
            "CIN: U72900MH2012PTC234985"
        )
    )

    style_range(
        ws,
        f"B{footer_row}:E{footer_row+4}",
        font=normal_font,
        alignment=left_alignment,
        left=True,
        right=True,
        top=True,
        bottom=True
    )    

    ws.merge_cells(
        start_row=footer_row,
        start_column=6,
        end_row=footer_row + 4,
        end_column=9
    )

    ws.cell(
        row=footer_row,
        column=6,
        value=(
            "for Encept Premedia Pvt. Ltd.\n\n\n\n"
            "Authorized Signatory"
        )
    )

    style_range(
        ws,
        f"F{footer_row}:I{footer_row+4}",
        font=normal_font,
        alignment=center_alignment,
        left=True,
        right=True,
        top=True,
        bottom=True
    )

    # =========================
    # WEBSITE FOOTER
    # =========================

    website_row = footer_row + 5

    ws.merge_cells(
        start_row=website_row,
        start_column=2,
        end_row=website_row,
        end_column=9
    )

    ws.cell(
        row=website_row,
        column=2,
        value="Visit us on www.enceptglobal.com"
    )

    style_range(
        ws,
        f"B{website_row}:I{website_row}",
        fill=header_fill,
        font=bold_font,
        alignment=center_alignment,
        left=True,
        right=True,
        top=True,
        bottom=True
    )

    
    # =========================
    # SIGNATURE IMAGE
    # =========================

    try:

        signature = Image(
            "VPsignature.jpg"
        )

        signature.width = 180
        signature.height = 45

        ws.add_image(
            signature,
            f"G{footer_row+1}"
        )

    except Exception as e:

        print(
            f"Signature insert failed: {e}"
        )


    # =========================
    # FILE NAME
    # =========================

    safe_brand = (

        str(brand)

        .replace("/", "_")
        .replace("\\\\", "_")
        .replace(":", "_")
        .replace("*", "_")
        .replace("?", "_")
        .replace('"', "_")
        .replace("<", "_")
        .replace(">", "_")
        .replace("|", "_")

    )

    safe_manager = (

        str(brand_manager)

        .replace("/", "_")
        .replace("\\\\", "_")
        .replace(":", "_")
        .replace("*", "_")
        .replace("?", "_")
        .replace('"', "_")
        .replace("<", "_")
        .replace(">", "_")
        .replace("|", "_")

    )

    file_name = (

        f"{sequence_str}_"
        f"EPPL_"
        f"{shortcode}_"
        f"{safe_brand}_"
        f"{current_month}_"
        f"{safe_manager}.xlsx"

    )

    wb.save(file_name)

    return file_name

# =========================
# FILE UPLOADS
# =========================

primary_file = st.file_uploader(
    "Upload Amt-Report Excel",
    type=["xlsx"]
)

secondary_file = st.file_uploader(
    "Upload Release-Report Excel",
    type=["xlsx"]
)

clients = []

master_df = None

# =========================
# PROCESS FILES
# =========================

if primary_file and secondary_file:

    try:

        primary_df = pd.read_excel(
            primary_file
        ).copy()

        primary_df.columns = (
            primary_df.columns
            .str.strip()
        )

        secondary_df = pd.read_excel(
            secondary_file
        ).copy()

        secondary_df.columns = (
            secondary_df.columns
            .str.strip()
        )

        column_mapping = {

            "Project_Task_ID": "Task_ID",
            "Client_Name": "Client_Name",
            "Project_Name": "Project_Name",
            "Brand_Name": "Brand_Name",
            "Variant_Name": "Variant_Name",
            "Package_Format": "Package_Format",
            "SKU_Name": "SKU_Name",
            "Release_To": "Revision_Type",
            "Release_Time": "AMT_Inititate_Time",
            "Project Type": "Project Type"

        }

        secondary_df = secondary_df.rename(
            columns=column_mapping
        )

        secondary_df["Task_Name"] = (
            "Release"
        )

        secondary_df = secondary_df.reindex(
            columns=primary_df.columns
        )

        master_df = pd.concat(

            [
                primary_df,
                secondary_df
            ],

            ignore_index=True

        )

        master_df.columns = (

            master_df.columns
            .str.strip()

        )

        clients = sorted(

            master_df["Client_Name"]
            .dropna()
            .unique()
            .tolist()

        )

        st.success(

            f"Master data created successfully "
            f"({len(master_df)} rows)"

        )

    except Exception as e:

        st.error(
            f"Error processing files: {e}"
        )

# =========================
# CLIENT SELECTOR
# =========================

client = st.selectbox(
    "Select Client",
    clients
)

# =========================
# DATE FILTERS
# =========================

col1, col2 = st.columns(2)

with col1:

    start_date = st.date_input(
        "Start Date"
    )

with col2:

    end_date = st.date_input(
        "End Date"
    )

# =========================
# GENERATE BUTTON
# =========================

if st.button("Generate Estimates"):

    try:

        if master_df is None:

            st.error(
                "Please upload both Excel files first."
            )

        else:

            master_df["AMT_Inititate_Time"] = (

                pd.to_datetime(

                    master_df[
                        "AMT_Inititate_Time"
                    ],

                    errors="coerce"

                ).dt.date

            )

            filtered_df = master_df[

                (
                    master_df["Client_Name"]
                    == client
                )

                &

                (
                    master_df[
                        "AMT_Inititate_Time"
                    ] >= start_date
                )

                &

                (
                    master_df[
                        "AMT_Inititate_Time"
                    ] <= end_date
                )

            ]

            st.success(
                f"{len(filtered_df)} records found"
            )

            brands = sorted(

                filtered_df["Brand_Name"]
                .dropna()
                .unique()
                .tolist()
            )

            generated_files = []
            for index, brand in enumerate(
                brands,
                start=1
            ):

                #st.subheader(brand)

                brand_df = filtered_df[

                    filtered_df["Brand_Name"]
                    == brand

                ]

                estimate_df = calculate_estimate(

                    brand_df,
                    client,
                    rate_master_df

                )

                #st.dataframe(
                #    estimate_df,
                #    use_container_width=True
                #)

                total_amount = (
                    estimate_df["Amount"]
                    .sum()
                )

                #st.success(

                #    f"Total Amount: "
                #    f"{total_amount:,.2f}"

                #)

                file_name = (
                    generate_estimate_excel(

                        index,
                        brand,
                        brand_df,
                        estimate_df,
                        client,
                        client_master_df,
                        brand_master_df

                    )
                )
                generated_files.append(file_name)

                #st.success(
                #    f"Excel generated: "
                #    f"{file_name}"
                #)
            st.success(
                f"{len(generated_files)} estimate files generated successfully."
            )
            zip_file_name = "All_Estimates.zip"

            with zipfile.ZipFile(zip_file_name, "w") as zipf:

                for file in generated_files:

                    zipf.write(file)

            with open(zip_file_name, "rb") as f:

                st.download_button(
                    label="Download All Estimates",
                    data=f,
                    file_name=zip_file_name,
                    mime="application/zip"
                    )

    except Exception as e:

        st.error(
            f"Processing Error: {e}"
        )

